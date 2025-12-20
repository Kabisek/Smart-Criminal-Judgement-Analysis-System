import json
import csv
import os
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from pathlib import Path
import io

app = Flask(__name__)

# Configure data storage
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
CSV_FILE = DATA_DIR / "conversions.csv"
JSON_STORAGE = DATA_DIR / "raw_inputs.json"

# Define CSV headers based on the sample JSON
CSV_HEADERS = [
    "source_file_name", "court_of_appeal_case_no", "high_court_case_no",
    "high_court_location", "judgment_date_coa", "judgment_date_hc",
    "date_of_offence", "judges", "offence_sections", "offence_category",
    "location_of_offence", "language_of_criminal", "brief_facts_summary",
    "evidence_type_primary", "evidence_type_secondary",
    "hc_offence_of_conviction_sections", "hc_sentence_type", "hc_fine_amount",
    "hc_compensation_amount", "hc_judgment_summary",
    "grounds_of_appeal_raw_text_summary", "grounds_of_appeal_structured_notes",
    "gnd_contradictions", "gnd_chain_of_custody", "gnd_illegal_search_or_raid",
    "gnd_wrong_identification", "gnd_dying_declaration_validity",
    "gnd_circumstantial_insufficient", "gnd_medical_inconsistency",
    "gnd_misdirection_on_law", "gnd_procedural_error", "gnd_new_evidence",
    "gnd_sentence_excessive_or_inadequate", "gnd_delay_prejudice",
    "gnd_judicial_bias_or_unfair_trial", "gnd_other",
    "gnd_other_description", "witness_evidence_analysis_summary",
    "num_prosecution_witnesses", "num_defence_witnesses", "eyewitness_present",
    "child_witness_present", "expert_evidence_present", "medical_evidence_strength",
    "forensic_evidence_present", "chain_of_custody_quality", "dying_declaration_present",
    "dying_declaration_reliability", "confession_present", "circumstantial_case",
    "probability_assessment_by_court", "defence_version_summary",
    "legal_errors_identified", "legal_errors_description",
    "procedural_defects_present", "procedural_defects_description",
    "directions_on_burden_of_proof_correct", "court_of_appeal_analysis_summary",
    "coa_final_outcome_class", "coa_conviction_status", "coa_sentence_type",
    "coa_fine_amount", "coa_compensation_amount", "brief_judgment_file_summary",
    "prosecution_counsel", "defence_counsel", "plea_of_accused", "appeal_type",
    "trial_method", "digital_evidence_present", "weapon_recovered", "weapon_type",
    "motive_established", "failure_to_cross_examine_material_points",
    "cause_of_death_exact_text", "hospital_treatment_details_present",
    "hc_errors_identified_by_coa", "hc_analysis_summary",
    "precedents_cited_list", "standard_of_review_applied",
    "sentencing_principles_applied", "final_charge_after_appeal",
    "order_on_retrial", "release_ordered", "reserved_for_future_use",
    "timestamp", "conversion_id"
]


def load_json_storage():
    """Load existing JSON storage or create new one."""
    if JSON_STORAGE.exists():
        with open(JSON_STORAGE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def save_json_storage(data):
    """Save JSON storage."""
    with open(JSON_STORAGE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def append_to_csv(row_data):
    """Append a row to the CSV file."""
    file_exists = CSV_FILE.exists()
    
    with open(CSV_FILE, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=CSV_HEADERS)
        if not file_exists:
            writer.writeheader()
        writer.writerow(row_data)


def validate_json(json_data):
    """Validate JSON structure."""
    try:
        if isinstance(json_data, str):
            json_obj = json.loads(json_data)
        else:
            json_obj = json_data
        
        # Check if it's a dictionary
        if not isinstance(json_obj, dict):
            return False, "JSON must be an object/dictionary"
        
        return True, json_obj
    except json.JSONDecodeError as e:
        return False, f"Invalid JSON: {str(e)}"


def process_json_row(json_obj):
    """Convert JSON object to CSV row."""
    row = {}
    
    for header in CSV_HEADERS:
        if header in ["timestamp", "conversion_id"]:
            continue
        
        value = json_obj.get(header)
        
        # Handle list fields - convert to pipe-separated string
        if isinstance(value, list):
            value = " | ".join(str(item) for item in value)
        
        row[header] = value if value is not None else ""
    
    # Add metadata
    row["timestamp"] = datetime.now().isoformat()
    row["conversion_id"] = f"CONV_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    return row


@app.route('/')
def index():
    """Serve the main page."""
    return render_template('index.html')


@app.route('/api/convert', methods=['POST'])
def convert_json():
    """Convert JSON to CSV row and save."""
    try:
        data = request.get_json()
        json_input = data.get('json_input')
        
        # Validate JSON
        is_valid, result = validate_json(json_input)
        if not is_valid:
            return jsonify({'success': False, 'error': result}), 400
        
        json_obj = result
        
        # Process and convert
        csv_row = process_json_row(json_obj)
        
        # Save to CSV
        append_to_csv(csv_row)
        
        # Save to JSON storage
        json_storage = load_json_storage()
        json_obj['timestamp'] = csv_row['timestamp']
        json_obj['conversion_id'] = csv_row['conversion_id']
        json_storage.append(json_obj)
        save_json_storage(json_storage)
        
        return jsonify({
            'success': True,
            'message': 'Data converted and saved successfully!',
            'conversion_id': csv_row['conversion_id'],
            'csv_row': csv_row
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/data', methods=['GET'])
def get_all_data():
    """Get all stored conversions."""
    try:
        json_storage = load_json_storage()
        
        # Count rows in CSV
        csv_rows = 0
        if CSV_FILE.exists():
            with open(CSV_FILE, 'r', encoding='utf-8') as f:
                csv_rows = sum(1 for line in f) - 1  # Exclude header
        
        return jsonify({
            'success': True,
            'total_entries': len(json_storage),
            'csv_rows': csv_rows,
            'data': json_storage
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download/csv', methods=['GET'])
def download_csv():
    """Download the CSV file."""
    try:
        if not CSV_FILE.exists():
            return jsonify({'error': 'No data available'}), 404
        
        return send_file(
            CSV_FILE,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'conversions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/json', methods=['GET'])
def download_json():
    """Download the JSON storage file."""
    try:
        json_storage = load_json_storage()
        
        if not json_storage:
            return jsonify({'error': 'No data available'}), 404
        
        json_bytes = json.dumps(json_storage, indent=2, ensure_ascii=False).encode('utf-8')
        
        return send_file(
            io.BytesIO(json_bytes),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'raw_inputs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/excel', methods=['GET'])
def download_excel():
    """Download as Excel format."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        if not CSV_FILE.exists():
            return jsonify({'error': 'No data available'}), 404
        
        # Read CSV and create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Write headers
        for col_idx, header in enumerate(CSV_HEADERS, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        with open(CSV_FILE, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row_idx, row in enumerate(reader, 2):
                for col_idx, header in enumerate(CSV_HEADERS, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'conversions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except ImportError:
        return jsonify({'error': 'Excel export requires openpyxl. Install with: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/delete/<conversion_id>', methods=['DELETE'])
def delete_entry(conversion_id):
    """Delete a specific entry."""
    try:
        # Delete from JSON storage
        json_storage = load_json_storage()
        json_storage = [item for item in json_storage if item.get('conversion_id') != conversion_id]
        save_json_storage(json_storage)
        
        # Delete from CSV (recreate without the entry)
        if CSV_FILE.exists():
            with open(CSV_FILE, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                rows = [row for row in reader if row.get('conversion_id') != conversion_id]
            
            with open(CSV_FILE, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=CSV_HEADERS)
                writer.writeheader()
                writer.writerows(rows)
        
        return jsonify({'success': True, 'message': 'Entry deleted successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get statistics about stored data."""
    try:
        json_storage = load_json_storage()
        
        stats = {
            'total_entries': len(json_storage),
            'csv_file_exists': CSV_FILE.exists(),
            'csv_file_size': CSV_FILE.stat().st_size if CSV_FILE.exists() else 0,
            'json_file_size': JSON_STORAGE.stat().st_size if JSON_STORAGE.exists() else 0,
            'latest_entry': json_storage[-1].get('timestamp') if json_storage else None
        }
        
        return jsonify({'success': True, 'stats': stats})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
